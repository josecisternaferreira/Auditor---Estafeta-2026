import streamlit as st
import pandas as pd
import io
import json
import unicodedata
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# Configuración de página
st.set_page_config(
    page_title="Auditor Estafeta Abril 2026",
    page_icon="📊",
    layout="wide"
)

DATA_PATH = Path(__file__).parent / 'data' / 'tarifas_2026.json'


def normalizar(texto):
    """Uppercase, sin tildes, sin guiones, para matchear nombres de ciudad."""
    texto = unicodedata.normalize('NFKD', str(texto))
    texto = ''.join(c for c in texto if not unicodedata.combining(c))
    return texto.upper().strip().replace('-', ' ')


@st.cache_data
def cargar_datos_tarifarios():
    """Carga tarifas (matriz completa 19 orígenes x ~305 destinos) y ramales
    desde data/tarifas_2026.json (extraído de la Tarifa Cuentas Corrientes
    CLIENTE VIP 2026 Actualizada Abril). Devuelve todo normalizado para
    matchear directamente contra el Origen/Destino del Estado de Pago."""
    with open(DATA_PATH, encoding='utf-8') as f:
        data = json.load(f)

    tarifas_norm = {
        normalizar(origen): {normalizar(destino): valores for destino, valores in destinos.items()}
        for origen, destinos in data['tarifas'].items()
    }

    ramales_norm = {}
    for seccion in data['ramales'].values():
        ciudad_base_norm = normalizar(seccion['ciudad_base'])
        for ciudad, cargo in seccion['destinos'].items():
            ramal_norm = normalizar(ciudad)
            if ramal_norm not in ramales_norm:
                ramales_norm[ramal_norm] = {'ciudad_base': ciudad_base_norm, 'cargo': cargo}

    return tarifas_norm, ramales_norm

def calcular_tarifa_base(peso_facturable, tarifas_ruta):
    """Calcula tarifa base según peso"""
    if peso_facturable <= 1:
        return tarifas_ruta.get('carta', 0), 'Carta (≤1kg)'
    elif peso_facturable <= 5:
        return tarifas_ruta.get('v5', 0), 'Hasta 5kg'
    elif peso_facturable <= 10:
        return tarifas_ruta.get('v10', 0), 'Hasta 10kg'
    elif peso_facturable <= 15:
        return tarifas_ruta.get('v15', 0), 'Hasta 15kg'
    elif peso_facturable <= 20:
        return tarifas_ruta.get('v20', 0), 'Hasta 20kg'
    elif peso_facturable <= 25:
        return tarifas_ruta.get('v25', 0), 'Hasta 25kg'
    elif peso_facturable <= 30:
        return tarifas_ruta.get('v30', 0), 'Hasta 30kg'
    elif peso_facturable <= 35:
        return tarifas_ruta.get('v35', 0), 'Hasta 35kg'
    elif peso_facturable <= 40:
        return tarifas_ruta.get('v40', 0), 'Hasta 40kg'
    elif peso_facturable <= 45:
        return tarifas_ruta.get('v45', 0), 'Hasta 45kg'
    elif peso_facturable <= 50:
        return tarifas_ruta.get('v50', 0), 'Hasta 50kg'
    else:
        base_50kg = tarifas_ruta.get('v50', 0)
        kg_adicionales = peso_facturable - 50
        
        # LÓGICA CORRECTA: comparar kg_adicionales (EXCESO) con 100, 500
        if kg_adicionales <= 100:
            tarifa_kg = tarifas_ruta.get('hasta100kg', 0)
        elif kg_adicionales <= 500:
            tarifa_kg = tarifas_ruta.get('hasta500kg', 0)
        else:
            tarifa_kg = tarifas_ruta.get('mas500kg', 0)
        
        total = base_50kg + (kg_adicionales * tarifa_kg)
        detalle = f'Base 50kg + {kg_adicionales:.1f}kg × ${tarifa_kg}'
        return total, detalle

def auditar_excel(df):
    """Audita el archivo de Estafeta"""
    tarifas, ramales = cargar_datos_tarifarios()
    resultados = []

    for idx, row in df.iterrows():
        try:
            orden = int(row['OrdFle'])
            origen_raw = str(row['Origen']).upper().strip()
            destino_raw = str(row['Destino']).upper().strip()

            peso_real = float(row['Kg.Bul']) if pd.notna(row['Kg.Bul']) else 0
            peso_vol = float(row['Kg.Vol.']) if pd.notna(row['Kg.Vol.']) else 0
            peso_fact = max(peso_real, peso_vol)
            cobrado_neto = float(row['Neto Fac']) if pd.notna(row['Neto Fac']) else 0

            origen = normalizar(origen_raw)
            destino = normalizar(destino_raw)

            estado = 'SIN_TARIFA'
            diferencia = None
            correcto_desc = None
            tramo_desc = ''
            es_ramal = False
            cargo_ramal = 0

            tarifas_ruta = tarifas.get(origen, {}).get(destino)

            if tarifas_ruta is None and destino in ramales:
                # Destino no tiene tarifa directa desde este origen: se trata
                # como ramal (tarifa de la ciudad base del ramal + cargo fijo)
                ramal_info = ramales[destino]
                tarifas_ruta = tarifas.get(origen, {}).get(ramal_info['ciudad_base'])
                if tarifas_ruta is not None:
                    cargo_ramal = ramal_info['cargo']
                    es_ramal = True

            if tarifas_ruta is not None:
                tarifa_base, tramo_desc = calcular_tarifa_base(peso_fact, tarifas_ruta)

                tarifa_total_sin_desc = tarifa_base + cargo_ramal
                correcto_desc = round(tarifa_total_sin_desc * 0.80, 2)

                diferencia = cobrado_neto - correcto_desc

                if abs(diferencia) < 10:
                    estado = 'CORRECTO'
                elif diferencia > 0:
                    estado = 'SOBRECOBRO'
                else:
                    estado = 'COBRO_MENOS'

            if es_ramal:
                tramo_desc = f"{tramo_desc} + Ramal {destino_raw}"

            resultados.append({
                'Orden': orden,
                'Fecha': row['Fecha'] if 'Fecha' in row else '',
                'Origen': origen_raw,
                'Destino': destino_raw,
                'Peso_Real_kg': peso_real,
                'Peso_Vol_kg': peso_vol,
                'Peso_Fact_kg': peso_fact,
                'Tramo': tramo_desc,
                'Tarifa_Correcta': correcto_desc,
                'Cobrado_Neto': cobrado_neto,
                'Diferencia': diferencia,
                'Estado': estado,
            })
        except Exception as e:
            st.warning(f"Error procesando fila {idx}: {str(e)}")
            continue
    
    return pd.DataFrame(resultados)

def generar_excel_auditoria(df_audit):
    """Genera Excel con formato"""
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_audit.to_excel(writer, index=False, sheet_name='Auditoría')
        
        workbook = writer.book
        worksheet = writer.sheets['Auditoría']
        
        # Colores
        color_header = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        color_ok = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        color_error = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        font_header = Font(bold=True, color='FFFFFF')
        
        # Formatear encabezados
        for cell in worksheet[1]:
            cell.fill = color_header
            cell.font = font_header
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Formatear filas según estado
        for idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=0):
            if idx < len(df_audit):
                estado = df_audit.iloc[idx]['Estado']
                estado_col = 12
                
                if estado == 'CORRECTO':
                    worksheet.cell(row=idx+2, column=estado_col).fill = color_ok
                elif estado == 'SOBRECOBRO':
                    worksheet.cell(row=idx+2, column=estado_col).fill = color_error
        
        # Ajustar anchos
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    return output

# ========== INTERFAZ STREAMLIT ==========
st.title("📊 Auditor Estafeta Abril 2026")
st.info("🔥 **Tarifas actualizadas por aumento de combustible - Vigentes desde Abril 2026**")
st.markdown("---")

# Información
with st.expander("ℹ️ Instrucciones de uso"):
    st.markdown("""
    ### Cómo usar el auditor:
    1. **Sube tu Estado de Pago** de Estafeta (archivo Excel)
    2. El sistema detectará automáticamente las columnas necesarias
    3. Haz clic en **AUDITAR**
    4. Descarga los resultados en Excel
    
    ### Qué hace el auditor:
    - ✅ Verifica cada envío contra la tarifa oficial Abril 2026 (con aumento combustible)
    - ✅ Aplica automáticamente el 20% de descuento corporativo
    - ✅ Detecta ramales (ej: Mejillones) y aplica cargos adicionales
    - ✅ Identifica sobrecobros y subcobros
    - ✅ Genera reporte descargable con % de desviación
    """)

# Upload
uploaded_file = st.file_uploader(
    "📁 Sube tu Estado de Pago (Excel)",
    type=['xlsx', 'xls'],
    help="Archivo Excel con las columnas: OrdFle, Origen, Destino, Kg.Bul, Kg.Vol., Neto Fac"
)

if uploaded_file:
    try:
        # Leer Excel
        df = pd.read_excel(uploaded_file, header=4)
        df = df[pd.notna(df['OrdFle'])].copy()
        df = df[df['Origen'].notna()].copy()
        
        st.success(f"✅ Archivo cargado: {len(df)} órdenes encontradas")
        
        # Vista previa
        with st.expander("👀 Vista previa de datos"):
            st.dataframe(df.head(10))
        
        # Botón auditar
        if st.button("🚀 AUDITAR", type="primary", use_container_width=True):
            with st.spinner("Auditando..."):
                df_audit = auditar_excel(df)
            
            st.success("✅ Auditoría completada")
            
            # Métricas
            col1, col2, col3, col4, col5 = st.columns(5)

            total = len(df_audit)
            correctos = len(df_audit[df_audit['Estado'] == 'CORRECTO'])
            sobrecobros = len(df_audit[df_audit['Estado'] == 'SOBRECOBRO'])
            cobros_menos = len(df_audit[df_audit['Estado'] == 'COBRO_MENOS'])
            sin_tarifa = len(df_audit[df_audit['Estado'] == 'SIN_TARIFA'])

            col1.metric("Total Órdenes", total)
            col2.metric("✅ Correctos", correctos)
            col3.metric("🔴 Sobrecobros", sobrecobros)
            col4.metric("🟢 Cobros Menos", cobros_menos)
            col5.metric("⚠️ Sin Tarifa", sin_tarifa)
            
            # Resumen financiero
            st.markdown("### 💰 Resumen Financiero")
            col1, col2, col3 = st.columns(3)
            
            total_cobrado = df_audit['Cobrado_Neto'].sum()
            total_correcto = df_audit[df_audit['Tarifa_Correcta'].notna()]['Tarifa_Correcta'].sum()
            diferencia_neta = df_audit[df_audit['Diferencia'].notna()]['Diferencia'].sum()
            
            # Calcular porcentajes de desviación
            if total_correcto > 0:
                porcentaje_diferencia_neta = (diferencia_neta / total_correcto) * 100
            else:
                porcentaje_diferencia_neta = 0
            
            col1.metric("Total Cobrado", f"${total_cobrado:,.0f}")
            col2.metric("Total Correcto", f"${total_correcto:,.0f}")
            
            if diferencia_neta > 0:
                col3.metric(
                    "Diferencia Neta", 
                    f"${diferencia_neta:,.0f}", 
                    delta=f"{porcentaje_diferencia_neta:+.2f}% Sobrecobro", 
                    delta_color="inverse"
                )
            elif diferencia_neta < 0:
                col3.metric(
                    "Diferencia Neta", 
                    f"${abs(diferencia_neta):,.0f}", 
                    delta=f"{abs(porcentaje_diferencia_neta):.2f}% A favor", 
                    delta_color="normal"
                )
            else:
                col3.metric("Diferencia Neta", "$0", delta="0.00% Perfecto")
            
            # Detalles adicionales de desviación
            if sobrecobros > 0 or cobros_menos > 0:
                st.markdown("### 📊 Detalle de Desviaciones")
                col1, col2, col3 = st.columns(3)
                
                # Sobrecobros
                sobrecobros_df = df_audit[df_audit['Estado'] == 'SOBRECOBRO']
                if len(sobrecobros_df) > 0:
                    total_sobrecobros = sobrecobros_df['Diferencia'].sum()
                    porcentaje_sobrecobros = (total_sobrecobros / total_correcto) * 100 if total_correcto > 0 else 0
                    col1.metric(
                        "🔴 Total Sobrecobros", 
                        f"${total_sobrecobros:,.0f}",
                        delta=f"+{porcentaje_sobrecobros:.2f}%"
                    )
                else:
                    col1.metric("🔴 Total Sobrecobros", "$0", delta="0.00%")
                
                # Cobros menos
                cobros_menos_df = df_audit[df_audit['Estado'] == 'COBRO_MENOS']
                if len(cobros_menos_df) > 0:
                    total_cobros_menos = abs(cobros_menos_df['Diferencia'].sum())
                    porcentaje_cobros_menos = (total_cobros_menos / total_correcto) * 100 if total_correcto > 0 else 0
                    col2.metric(
                        "🟢 Total Cobros Menos", 
                        f"${total_cobros_menos:,.0f}",
                        delta=f"-{porcentaje_cobros_menos:.2f}%"
                    )
                else:
                    col2.metric("🟢 Total Cobros Menos", "$0", delta="0.00%")
                
                # Neto
                col3.metric(
                    "⚖️ Balance Neto",
                    f"${abs(diferencia_neta):,.0f}",
                    delta=f"{porcentaje_diferencia_neta:+.2f}%"
                )
            
            # Detalle de órdenes sin tarifa (ruta no encontrada en la tabla)
            if sin_tarifa > 0:
                st.markdown("### ⚠️ Órdenes Sin Tarifa (no encontradas en la tabla)")
                st.warning(f"Se encontraron **{sin_tarifa}** órdenes cuya ruta Origen-Destino no está en la tabla de tarifas. Deben revisarse manualmente.")
                sin_tarifa_df = df_audit[df_audit['Estado'] == 'SIN_TARIFA'][['Orden', 'Origen', 'Destino', 'Peso_Fact_kg', 'Cobrado_Neto']]
                st.dataframe(sin_tarifa_df, use_container_width=True)

            # Detalles de sobrecobros
            if sobrecobros > 0:
                st.markdown("### 🔴 Sobrecobros Detectados")
                sobrecobros_df = df_audit[df_audit['Estado'] == 'SOBRECOBRO'][['Orden', 'Origen', 'Destino', 'Peso_Fact_kg', 'Cobrado_Neto', 'Tarifa_Correcta', 'Diferencia']]
                st.dataframe(sobrecobros_df, use_container_width=True)
                
                total_reclamable = sobrecobros_df['Diferencia'].sum()
                st.info(f"💵 **Total reclamable: ${total_reclamable:,.0f}**")

            # Detalles de cobros menos
            if cobros_menos > 0:
                st.markdown("### 🟢 Cobros Menos Detectados")
                cobros_menos_df = df_audit[df_audit['Estado'] == 'COBRO_MENOS'][['Orden', 'Origen', 'Destino', 'Peso_Fact_kg', 'Cobrado_Neto', 'Tarifa_Correcta', 'Diferencia']]
                st.dataframe(cobros_menos_df, use_container_width=True)

                total_a_favor = abs(cobros_menos_df['Diferencia'].sum())
                st.info(f"💵 **Total a favor: ${total_a_favor:,.0f}**")

            # Tabla completa
            with st.expander("📋 Ver auditoría completa"):
                st.dataframe(df_audit, use_container_width=True)
            
            # Descarga
            excel_output = generar_excel_auditoria(df_audit)
            
            st.download_button(
                label="📥 Descargar Auditoría (Excel)",
                data=excel_output,
                file_name=f"AUDITORIA_ESTAFETA_ABRIL_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
    
    except Exception as e:
        st.error(f"❌ Error al procesar archivo: {str(e)}")
        st.info("Asegúrate de que el archivo tenga las columnas: OrdFle, Origen, Destino, Kg.Bul, Kg.Vol., Neto Fac")

# Footer
st.markdown("---")
st.markdown("""
<div style='text-align: center; color: #666;'>
    <small>Auditor Estafeta Abril 2026 | Tarifas actualizadas con aumento combustible | 20% descuento corporativo</small>
</div>
""", unsafe_allow_html=True)
